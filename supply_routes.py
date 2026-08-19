"""Заявка в магазин и приход от него: Excel туда, Excel обратно.

Цикл, ради которого всё:

    1. Программа считает, чего не хватает до нормы  (stock_routes.handle_order)
    2. Выгружаем это в .xlsx и отправляем в магазин  → handle_export
    3. Магазин проставляет, что у него реально есть, и присылает файл назад
    4. Загружаем файл обратно                        → handle_import
    5. Получается поставка: что и сколько мы правда заберём
    6. Водитель видит её у себя и сканирует бутылки при получении

Ключ к 3–4 шагам — колонка с кодом позиции. Магазин правит количества и может
переставить или удалить строки; по названию сопоставлять нельзя, потому что
«Absolut 1 ltr» у них в файле легко станет «ABSOLUT BLUE 1L». Код едет в файле
и возвращается неизменным, а название рядом — чтобы человеку было понятно.

Лист один и сразу по точкам: «сколько всего» — это сумма по строке, и хранить
её отдельным листом значило держать два числа об одном и том же. Заодно из
файла возвращается развозка — какая бутылка в какое здание, а не общий ворох.
"""
import io
import logging
import re
import zipfile
from datetime import datetime, timezone

from aiohttp import web

import db
from owner_auth import require_owner, CORS_HEADERS
from config_offices import OFFICE_IDS, OFFICE_NAMES, OFFICE_CODES
from config_stock_order import order_key      # порядок обхода полок, как в их таблице

log = logging.getLogger("supply")

# Файл уходит в магазин — он на английском целиком. Заголовки читает их
# сотрудник, и «Подтверждено» ему ничего не говорит.
CODE_COL = "Code"          # служебная колонка, по ней идёт возврат
TOTAL_COL = "Total"          # считается формулой, а не нами

# Ноль показываем пустой клеткой. В файле все позиции каталога, и лист, залитый
# нулями, читать невозможно: глаз ищет числа, а видит шум.
ZERO_BLANK = "0;\\-0;;@"
# Пустые строки под итогом. Итог — последняя строка листа, и на телефоне
# просмотрщик прижимает её к самому низу экрана: она наполовину уходит под
# край и её приходится выкручивать пальцем. Несколько пустых строк снизу
# ничего не стоят, а итог перестаёт липнуть к краю.
TAIL_ROWS = 8
SHEET_MAIN = "Order"

# Названия точек у нас записаны кириллицей, хотя районы английские.
DIST_EN = {"jvc": "JVC", "bbay": "Business Bay", "silicon": "Silicon Oasis",
           "alguses": "Al Qusais", "tecom": "Tecom"}


def _with_cached_values(raw: bytes, calc: dict) -> bytes:
    """Дописать формулам посчитанный результат.

    Книга, собранная программой, содержит формулу и ничего больше: значение
    появится, когда файл откроют в Excel и он пересчитает. Но открывают его
    сперва не в Excel — в предпросмотре телефона, в телеграме, в почте. Там
    никто ничего не считает, и вся колонка Total показывает нули: файл выглядит
    сломанным раньше, чем до него дойдут руки.

    Поэтому кладём рядом с формулой её результат. Excel всё равно пересчитает
    сам (fullCalcOnLoad), а просмотрщик покажет то, что мы посчитали.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw)) as src, \
         zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                xml = data.decode("utf-8")

                def put(m):
                    v = calc.get(m.group(2))
                    if v is None:
                        return m.group(0)
                    # Формула вернёт пустоту — её и кладём, иначе просмотрщик
                    # покажет ноль там, где мы ничего не просили.
                    if v == "":
                        tag = m.group(1)
                        if ' t="' not in tag:
                            tag = tag[:-1] + ' t="str">'
                        return f"{tag}{m.group(3)}<v></v></c>"
                    return f"{m.group(1)}{m.group(3)}<v>{v}</v></c>"

                # Пустой <v/> после формулы openpyxl ставит сам — в него и
                # смотрит просмотрщик, показывая ноль.
                xml = re.sub(
                    r'(<c r="([A-Z]+\d+)"[^>]*>)(<f>.*?</f>)(?:<v\s*/>|<v>[^<]*</v>)?</c>',
                    put, xml)
                data = xml.encode("utf-8")
            dst.writestr(item, data)
    return buf.getvalue()


def _catalog_by_id():
    from operator_routes import _catalog_by_id as f
    return f()


async def _order_rows(day):
    """Строки заявки — тем же расчётом, что и на экране «Закуп»."""
    import stock_routes
    from aiohttp.test_utils import make_mocked_request
    # Логика заявки живёт в stock_routes и там же должна остаться: два расчёта
    # «сколько докупить» рано или поздно разойдутся, и никто не заметит.
    return await stock_routes.order_rows(day)


@require_owner
async def handle_export(request):
    """Заявка файлом по ссылке — для браузера."""
    raw, name = await _build_book((request.query.get("day") or "").strip())
    return web.Response(
        body=raw,
        headers={**CORS_HEADERS,
                 "Content-Type": "application/vnd.openxmlformats-officedocument."
                                 "spreadsheetml.sheet",
                 "Content-Disposition": f'attachment; filename="{name}"'})


@require_owner
async def handle_send(request):
    """Прислать заявку файлом в телеграм — так её и отправляют дальше.

    В мини-приложении скачать файл некуда: браузера у него нет, а «загрузки» на
    телефоне ещё поискать. Зато переслать документ из своего же чата магазину —
    одно движение, поэтому бот кладёт файл владельцу в переписку."""
    from api_server import _aiohttp
    from owner_routes import OWNER_BOT_TOKEN
    if not OWNER_BOT_TOKEN:
        return web.json_response({"error": "no_bot"}, status=500, headers=CORS_HEADERS)
    raw, name = await _build_book((request.query.get("day") or "").strip())
    form = _aiohttp.FormData()
    form.add_field("chat_id", str(request.get("owner_id") or 0))
    form.add_field("caption", f"Заявка в магазин · {name[15:-5]}")
    form.add_field("document", raw, filename=name,
                   content_type="application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet")
    url = f"https://api.telegram.org/bot{OWNER_BOT_TOKEN}/sendDocument"
    to = _aiohttp.ClientTimeout(total=30)
    try:
        async with _aiohttp.ClientSession(timeout=to) as sess:
            async with sess.post(url, data=form) as r:
                res = await r.json()
    except Exception as e:
        log.error(f"[supply] отправка файла: {e}")
        return web.json_response({"error": "send_failed"}, status=502, headers=CORS_HEADERS)
    if not res.get("ok"):
        log.error(f"[supply] телеграм отказал: {res.get('description')}")
        return web.json_response({"error": "telegram", "detail": res.get("description")},
                                 status=502, headers=CORS_HEADERS)
    return web.json_response({"ok": True, "file": name}, headers=CORS_HEADERS)


async def _build_book(day: str):
    """Заявка в .xlsx — тот файл, который уходит в магазин.

    Вид листа задан владельцем: он один раз разложил его так, как удобно
    человеку в магазине, и дальше файл собирается по этому образцу. Отсюда
    пустая колонка слева, нумерация строк, чёрная шапка и красная полоса с
    просьбой — их видно раньше, чем начинают читать.

    Один лист и сразу в разрезе точек. Отдельная «просто заявка» была лишней:
    магазин всё равно смотрит в неё, а нам потом нужно знать, куда развозить —
    и два списка приходилось сверять глазами.

    Позиции все, включая те, что сейчас не нужны. Магазину так проще: это
    привычный ему прайс, в котором он правит числа, а не список из пятидесяти
    строк, где не найти то, что он хочет предложить сверх заказа.

    Total в строке — формула. Магазин правит числа по точкам, и переписанный
    руками итог разошёлся бы с ними в первый же раз."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    data = await _order_rows(day)
    # В файл едет только то, что просим. Позиция, которой не нужно ни одной
    # бутылки, — это пустая строка среди сотни таких же: магазин листает их
    # глазами, ищет, где же цифры, и однажды пропустит настоящую. Понадобится
    # в следующий раз хоть одна — строка вернётся сама, файл собирается заново
    # каждую смену.
    #
    # Цена решения: раньше магазин мог дописать количество в любую строку
    # каталога и предложить то, чего мы не просили. Теперь такой строки в файле
    # нет — вписывать придётся руками, с кодом.
    src = [r for r in (data.get("all_rows") or data["rows"]) if r["need_total"] > 0]
    # Порядок как в их таблице: заявку собирают, идя вдоль полок, и список,
    # отсортированный по количеству, заставляет бегать по залу кругами.
    rows = sorted(src, key=lambda r: order_key(r["id"]))
    dist = [d["id"] for d in data["districts"]]

    wb = Workbook()
    # Excel пересчитывает книгу при открытии: подставленные нами значения —
    # для просмотрщиков, а живая правка чисел по точкам должна менять итог.
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = SHEET_MAIN
    ws.sheet_view.showGridLines = False        # рамки рисуем сами, сетка мешает

    white = Font(bold=True, size=11, color="FFFFFFFF")
    bold = Font(bold=True, size=11)
    head_fill = PatternFill("solid", fgColor="FF1F2A37")
    num_fill = PatternFill("solid", fgColor="FF000000")     # колонка «№»
    ask = PatternFill("solid", fgColor="FFFFF3CD")          # что правит магазин
    # Каждая вторая строка обесцвечена целиком: глаз ведёт линию через десяток
    # колонок и без полосы соскакивает на соседнюю.
    band = PatternFill("solid", fgColor="FFEEECE1")
    sum_fill = PatternFill("solid", fgColor="FFFFFF00")     # итоги
    warn_fill = PatternFill("solid", fgColor="FFFF0000")    # просьба к магазину
    thin = Side(style="thin")
    med = Side(style="medium")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    mid = Alignment(horizontal="center", vertical="center")

    # Первая колонка пустая и широкая — поле, за которое лист приятно держать
    # глазами. Всё остальное начинается с B.
    N, C, I, D0 = 2, 3, 4, 5                   # №, Code, Item, первая точка
    LAST = D0 + len(dist) - 1                  # последняя точка
    TOT = LAST + 1                             # Total

    ws.cell(row=1, column=N, value=f"AMBAR · purchase order · {data['day']}")
    ws.cell(row=1, column=N).font = Font(bold=True, size=16)
    ws.cell(row=1, column=N).border = Border(bottom=thin)
    ws.merge_cells(start_row=1, start_column=N, end_row=1, end_column=I)
    ws.row_dimensions[1].height = 21.6

    ws.cell(row=2, column=N,
            value="Please correct the quantities you can supply and send the file back. "
                  "The Total column adds up by itself. Do not change the Code column.")
    ws.merge_cells(start_row=2, start_column=N, end_row=2, end_column=TOT)
    for col in range(N, TOT + 1):
        c = ws.cell(row=2, column=col)
        c.fill = warn_fill; c.font = Font(bold=True, size=12)
        c.alignment = mid
        c.border = Border(left=med, right=med, top=med, bottom=med)
    ws.row_dimensions[2].height = 33.6

    head = ["№", CODE_COL, "Item"] + \
           [f"{OFFICE_CODES.get(o,'')} {DIST_EN.get(o, OFFICE_NAMES.get(o,o))}" for o in dist] + \
           [TOTAL_COL]
    for i, title in enumerate(head):
        c = ws.cell(row=3, column=N + i, value=title)
        c.alignment = mid; c.border = box
        if N + i == N:      c.fill = num_fill;  c.font = white
        elif N + i == TOT:  c.fill = sum_fill;  c.font = bold
        else:               c.fill = head_fill; c.font = white
    ws.row_dimensions[3].height = 15

    first = 4
    calc = {}                      # ссылка ячейки → посчитанный итог
    for n, r in enumerate(rows, 1):
        i = first + n - 1
        ws.cell(row=i, column=N, value=n).alignment = mid
        ws.cell(row=i, column=C, value=r["id"]).alignment = mid
        nm = ws.cell(row=i, column=I, value=r["name"])
        nm.font = bold
        nm.alignment = Alignment(horizontal="left", vertical="center")
        stripe = (n % 2 == 0)          # полоса через строку — вести взгляд вдоль
        for k, o in enumerate(dist):
            # Ноль не пишем вовсе: на листе из ста двадцати строк колонка нулей
            # мешает увидеть те несколько чисел, ради которых заявку и читают.
            need = (r["cells"].get(o) or {}).get("need", 0) or None
            c = ws.cell(row=i, column=D0 + k, value=need)
            c.fill = band if stripe else ask
            c.alignment = mid; c.number_format = ZERO_BLANK
        # Итог строки считает сам файл: магазин правит числа по точкам, и
        # переписанная руками сумма разошлась бы с ними на первой же правке.
        rng = f"{get_column_letter(D0)}{i}:{get_column_letter(LAST)}{i}"
        t = ws.cell(row=i, column=TOT, value=f'=IF(SUM({rng})=0,"",SUM({rng}))')
        calc[f"{get_column_letter(TOT)}{i}"] = r["need_total"] or ""
        t.fill = band if stripe else sum_fill
        t.font = bold
        t.alignment = mid; t.number_format = ZERO_BLANK
        for col in range(N, TOT + 1):
            c = ws.cell(row=i, column=col)
            c.border = box
            if stripe and col < D0:
                c.fill = band
        ws.row_dimensions[i].height = 17.4

    last = first + len(rows) - 1
    i = last + 1
    ws.cell(row=i, column=N, value="TOTAL")
    ws.merge_cells(start_row=i, start_column=N, end_row=i, end_column=I)
    for col in range(N, TOT + 1):
        c = ws.cell(row=i, column=col)
        if col >= D0:
            L = get_column_letter(col)
            c.value = f'=IF(SUM({L}{first}:{L}{last})=0,"",SUM({L}{first}:{L}{last}))'
            c.number_format = ZERO_BLANK
            calc[f"{L}{i}"] = (sum(r["need_total"] for r in rows) if col == TOT
                               else sum((r["cells"].get(dist[col - D0]) or {}).get("need", 0)
                                        for r in rows)) or ""
        c.fill = sum_fill; c.font = bold; c.alignment = mid; c.border = box

    _tail(ws, i, N)

    ws.column_dimensions["A"].width = 29.55                 # пустое поле слева
    ws.column_dimensions[get_column_letter(N)].width = 7.11
    ws.column_dimensions[get_column_letter(C)].width = 10
    ws.column_dimensions[get_column_letter(I)].width = 40
    for col in range(D0, LAST + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions[get_column_letter(TOT)].width = 18.44
    # Держим на виду номер, код и название: магазин листает вправо по точкам и
    # без этого перестаёт понимать, в какой он строке.
    ws.freeze_panes = f"{get_column_letter(D0)}{first}"

    # Снимок того, что просили: в файле все позиции каталога, и вернувшийся ноль
    # сам по себе не отличает «магазин отказал» от «мы и не заказывали».
    try:
        await db.zayavka_save(
            data["day"],
            {r["id"]: r["need_total"] for r in rows if r["need_total"]},
            by={r["id"]: {o: c["need"] for o, c in r["cells"].items() if c["need"]}
                for r in rows if r["need_total"]})
    except Exception as e:
        log.warning(f"[supply] снимок заявки: {e}")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    raw = _with_cached_values(buf.read(), calc)
    name = f"AMBAR-zayavka-{data['day']}.xlsx"
    log.info(f"[supply] выгрузка заявки {data['day']}: {len(rows)} позиций, "
             f"из них с потребностью {sum(1 for r in rows if r['need_total'])}")
    return raw, name


def _tail(ws, last_row: int, col: int):
    """Пустые строки под итогом, чтобы он не упирался в край экрана.

    Строку в файле создаёт ячейка, а не высота: пустое значение openpyxl не
    записывает вовсе, и строгий просмотрщик считает строку несуществующей —
    рисует лист ровно до итога. Пустая строка «» даёт настоящую ячейку без
    единого символа внутри: и строка есть, и в ней ничего нет."""
    for r in range(last_row + 1, last_row + 1 + TAIL_ROWS):
        ws.cell(row=r, column=col, value="")
        ws.row_dimensions[r].height = 17.4


def _num(v):
    """Число из ячейки. Магазин пишет «12 шт», «12,0» и просто 12."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", ".").strip()
    if s.startswith("="):          # формула, которую Excel не пересчитал
        return None
    digits = "".join(ch for ch in s if ch.isdigit() or ch == ".")
    if not digits:
        return None
    try:
        return int(float(digits))
    except ValueError:
        return None


def _district_cols(cols: dict) -> dict:
    """Колонки точек по заголовкам: «B2 Business Bay» → офис.

    По коду, а не по названию: код мы ставим сами и он короткий, а название
    магазин может перевести или сократить."""
    out = {}
    for title, col in cols.items():
        head = str(title).strip()
        for oid in OFFICE_IDS:
            code = OFFICE_CODES.get(oid, "")
            if code and (head == code or head.startswith(code + " ")):
                out[oid] = col
                break
    return out


@require_owner
async def handle_import(request):
    """Ответ магазина: читаем файл и получаем поставку.

    Количество берём суммой по точкам, а не из Total: Total — формула, и если
    магазин правил файл в Google Sheets или в редакторе попроще, в ячейке
    приедет либо старое значение, либо сам текст формулы.

    Строка с нулями — это либо отказ, либо позиция, которую мы и не просили.
    Различаем по снимку заявки: без него пришлось бы либо терять отказы, либо
    записывать в них весь каталог."""
    from openpyxl import load_workbook

    reader = await request.multipart()
    field = await reader.next()
    if field is None or field.name != "file":
        return web.json_response({"error": "file_required"}, status=400, headers=CORS_HEADERS)
    raw = await field.read(decode=False)
    if not raw:
        return web.json_response({"error": "empty_file"}, status=400, headers=CORS_HEADERS)

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return web.json_response({"error": "not_xlsx", "detail": str(e)[:120]},
                                 status=400, headers=CORS_HEADERS)
    ws = wb[SHEET_MAIN] if SHEET_MAIN in wb.sheetnames else wb.worksheets[0]

    # Ищем строку заголовков, а не полагаемся на номер: магазин мог вставить
    # сверху свою шапку с логотипом, и жёсткая «третья строка» развалилась бы.
    hdr_row, cols = None, {}
    for row in ws.iter_rows(min_row=1, max_row=15):
        vals = {str(c.value).strip(): c.column for c in row if c.value}
        if CODE_COL in vals:
            hdr_row, cols = row[0].row, vals
            break
    if not hdr_row:
        return web.json_response({"error": "no_code_column", "need": CODE_COL},
                                 status=400, headers=CORS_HEADERS)

    dcols = _district_cols(cols)
    if not dcols:
        return web.json_response({"error": "no_district_columns"},
                                 status=400, headers=CORS_HEADERS)

    asked_map, asked_full, day = {}, {}, ""
    try:
        snap = await db.zayavka_last_full()
        asked_map = snap.get("asked") or {}
        asked_full = snap.get("by") or {}
        day = snap.get("_id") or ""
    except Exception as e:
        log.warning(f"[supply] снимок заявки не прочитан: {e}")

    cat = _catalog_by_id()
    # Разница считается здесь и один раз. Считать её потом, на экране, значит
    # пересчитывать при каждом открытии по снимку заявки, который к тому
    # времени уже сменится следующим днём.
    items, unknown, dropped, short, extra = [], [], [], [], []
    for row in ws.iter_rows(min_row=hdr_row + 1):
        code = row[cols[CODE_COL] - 1].value
        if not code:
            continue
        pid = str(code).strip()
        p = cat.get(pid)
        if not p:
            unknown.append(pid); continue
        by_district = {}
        for oid, col in dcols.items():
            n = _num(row[col - 1].value) or 0
            if n > 0:
                by_district[oid] = n
        qty = sum(by_district.values())
        asked = int(asked_map.get(pid) or 0)
        asked_by = {k: int(v) for k, v in (asked_full.get(pid) or {}).items()}
        name = p.get("name", "")
        if qty <= 0:
            # Отказом считаем только то, что правда просили: иначе в отказы
            # попал бы весь каталог, и водитель не нашёл бы в нём смысла.
            if asked:
                dropped.append({"id": pid, "name": name, "asked": asked,
                                "by_district": asked_by})
            continue
        if not asked:
            # Магазин дописал позицию, которую мы не просили. Это не ошибка —
            # так он предлагает то, что у него есть. Но это незапланированные
            # деньги, и увидеть их надо отдельно, а не в общем списке.
            extra.append({"id": pid, "name": name, "qty": qty,
                          "by_district": dict(by_district)})
        elif qty < asked:
            gap_by = {o: max(0, int(v) - int(by_district.get(o) or 0))
                      for o, v in asked_by.items()}
            short.append({"id": pid, "name": name, "asked": asked, "qty": qty,
                          "gap": asked - qty,
                          "by_district": {o: g for o, g in gap_by.items() if g}})
        items.append({"id": pid, "name": name,
                      "asked": asked, "qty": qty, "scanned": 0,
                      "by_district": by_district,
                      # Принято по районам считается отдельно от подтверждённого:
                      # районы забирают разные люди в разное время.
                      "got": {o: 0 for o in by_district}})

    if not items:
        return web.json_response({"error": "nothing_confirmed", "dropped": len(dropped)},
                                 status=400, headers=CORS_HEADERS)

    now = datetime.now(timezone.utc)
    sid = "S" + now.strftime("%y%m%d-%H%M%S")
    # Задача на район появляется сразу и пустой: водитель должен увидеть её в
    # своём приложении в ту же секунду, а не после того, как кто-то нажмёт
    # «отправить водителям». Лишний шаг здесь — это забытый шаг.
    tasks = {}
    for it in items:
        for oid, n in it["by_district"].items():
            t = tasks.setdefault(oid, {"qty": 0, "positions": 0, "scanned": 0,
                                       "driver": "", "driver_id": 0,
                                       "claimed_at": None, "started_at": None,
                                       "done_at": None, "last_at": None,
                                       "undo": 0, "note": "", "gaps": [], "flags": []})
            t["qty"] += n; t["positions"] += 1
    doc = {"_id": sid, "at": now, "status": "open", "day": day,
           "by": request.get("owner_id") or 0,
           "items": items, "dropped": dropped, "short": short, "extra": extra,
           "unknown": unknown, "tasks": tasks,
           "total_qty": sum(i["qty"] for i in items),
           "asked_qty": sum(int(v) for v in asked_map.values()),
           "gap_qty": sum(d["asked"] for d in dropped) + sum(s["gap"] for s in short)}
    await db.supply_save(doc)
    log.info(f"[supply] поставка {sid}: {len(items)} позиций, "
             f"{doc['total_qty']} бутылок, отказов {len(dropped)}, "
             f"урезано {len(short)}, сверх заказа {len(extra)}")
    return web.json_response({"ok": True, "supply_id": sid,
                              "items": len(items), "total_qty": doc["total_qty"],
                              "asked_qty": doc["asked_qty"], "gap_qty": doc["gap_qty"],
                              "dropped": dropped, "short": short, "extra": extra,
                              "unknown": unknown,
                              "tasks": [{"district": o, "code": OFFICE_CODES.get(o, ""),
                                         "name": OFFICE_NAMES.get(o, o), **t}
                                        for o, t in tasks.items()]},
                             headers=CORS_HEADERS)


@require_owner
async def handle_list(request):
    """Поставки: что в работе и что уже забрали."""
    rows = await db.supply_list(limit=30)
    return web.json_response({"supplies": rows}, headers=CORS_HEADERS,
                             dumps=lambda o: __import__("json").dumps(o, default=str))


# ════════════════════════════════════════════════════════════════════════════
#  ПРИЁМКА
# ════════════════════════════════════════════════════════════════════════════
# Поставку забирают по районам: у каждого своя машина и свой водитель. Задача =
# «поставка × район», её берут, по ней сканируют, её закрывают.
#
# Что здесь на самом деле защищается
# ----------------------------------
# Приём — единственное место, где остаток растёт со слов человека. Всё
# остальное в системе считается от заказов, а тут водитель says «привёз 23» — и
# склад ему верит. Поэтому:
#
#   • верит не словам, а кодам. Одна бутылка = один код = одна запись, второй
#     раз тот же код не вставить (он _id в реестре);
#   • строка задачи закрывается ровно на подтверждённом количестве. 24-ю
#     бутылку по строке из 23 принять нельзя — не «не нужно», а физически;
#   • недобор — обычный, дешёвый и честный путь. Если сказать «магазин дал 19
#     из 23» проще, чем отсканировать четыре лишних кода, то накрутка теряет
#     смысл. Это главная защита, остальные — вспомогательные;
#   • всё, что нельзя запретить, записывается и показывается старшему: темп
#     сканирования, отмены, коды непохожей формы, окно приёмки.
#
# Чего здесь нет и не будет: координат водителя. Ни при старте, ни при скане.
# Приёмку это не проверяет, а историю передвижений человека создаёт навсегда.

# Быстрее этого бутылки из коробки не достают: рука не успевает. Отдельный
# быстрый скан — случайность (две бутылки рядом), пять и больше — лист кодов.
FAST_MS = 1200
FAST_LIMIT = 5

# Отменить можно только что записанное — и не бесконечно. Без потолка отмена
# превращается в лазейку: отсканировал двадцать три, закрыл задачу, вернул три.
UNDO_SEC = 90
UNDO_MAX = 3


def _sig(code: str) -> str:
    """Форма кода: длина и набор классов символов.

    Коды на бутылках одной позиции приходят из одного источника и выглядят
    одинаково. Код, не похожий ни на один из уже записанных, — повод показать
    его старшему: чаще всего это чужая наклейка или картинка с экрана."""
    c = str(code or "")
    cls = "".join(sorted({("d" if ch.isdigit() else "a" if ch.islower()
                           else "A" if ch.isupper() else "-") for ch in c}))
    return f"{len(c) // 4}:{cls}"


_SIGS = {}          # product_id → (когда, {формы})


async def _known_sigs(pid: str) -> set:
    import time as _t
    hit = _SIGS.get(pid)
    if hit and _t.time() - hit[0] < 300:
        return hit[1]
    rows = await db.qr_list(pid, "", limit=60)
    sigs = {_sig(r.get("code") or "") for r in rows}
    _SIGS[pid] = (_t.time(), sigs)
    return sigs


def _task_view(sid: str, sup: dict, oid: str, task: dict, me: str = "") -> dict:
    """Задача глазами водителя: что забрать и сколько уже принято."""
    lines = []
    for it in sup.get("items") or []:
        need = int((it.get("by_district") or {}).get(oid) or 0)
        if not need:
            continue
        got = int((it.get("got") or {}).get(oid) or 0)
        lines.append({"id": it["id"], "name": it.get("name", ""),
                      "need": need, "got": got, "left": max(0, need - got)})
    # Недобранное — вверх: закрытые строки водителю больше не нужны, а искать
    # среди них следующую бутылку он будет каждый раз.
    lines.sort(key=lambda l: (l["left"] == 0, l["name"]))
    need = sum(l["need"] for l in lines)
    got = sum(l["got"] for l in lines)
    return {
        "supply_id": sid,
        "at": str(sup.get("at") or ""),
        "day": sup.get("day") or "",
        "district": oid,
        "district_code": OFFICE_CODES.get(oid, ""),
        "district_name": OFFICE_NAMES.get(oid, oid),
        "driver": task.get("driver") or "",
        "mine": bool(me and task.get("driver") == me),
        "claimed_at": str(task.get("claimed_at") or ""),
        "started_at": str(task.get("started_at") or ""),
        "done_at": str(task.get("done_at") or ""),
        "note": task.get("note") or "",
        "gaps": task.get("gaps") or [],
        "need": need, "got": got, "left": max(0, need - got),
        "positions": len(lines),
        "lines": lines,
    }


async def tasks_for_driver(me: str, district: str) -> dict:
    """Что водитель может взять и что уже взял.

    Свой район — сверху и всегда: это его товар и его полка. Чужие показываем
    ниже и не прячем: в магазин едет одна машина, и если водитель B1 всё равно
    там, забрать заодно ящик для B3 дешевле второго рейса."""
    mine, free, other = [], [], []
    for sup in await db.supplies_with_open_tasks(limit=8):
        sid = sup.get("_id")
        for oid, task in (sup.get("tasks") or {}).items():
            if task.get("done_at"):
                continue
            v = _task_view(sid, sup, oid, task, me)
            v["home"] = (oid == district)
            if v["mine"]:
                mine.append(v)
            elif v["driver"]:
                other.append(v)
            elif oid == district:
                free.insert(0, v)
            else:
                free.append(v)
    return {"mine": mine, "free": free, "taken": other}


async def task_scan(sid: str, oid: str, pid: str, code: str, me: str,
                    tg_id: int, at_dev: str = "") -> dict:
    """Принять одну бутылку. Возвращает исход, а не «ок» — их несколько.

    Порядок важен: сначала занимаем место в задаче, потом пишем бутылку в
    реестр. Наоборот — и упавшая на полпути запись оставила бы бутылку в
    остатке, но не в поставке: строка никогда бы не закрылась."""
    sup = await db.supply_get(sid)
    if not sup or sup.get("status") != "open":
        return {"ok": False, "verdict": "no_supply"}
    task = (sup.get("tasks") or {}).get(oid) or {}
    if task.get("driver") != me:
        return {"ok": False, "verdict": "not_mine", "driver": task.get("driver") or ""}
    if task.get("done_at"):
        return {"ok": False, "verdict": "closed"}

    item = next((i for i in (sup.get("items") or []) if i["id"] == pid), None)
    if not item:
        return {"ok": False, "verdict": "not_in_supply"}
    need = int((item.get("by_district") or {}).get(oid) or 0)
    got = int((item.get("got") or {}).get(oid) or 0)
    if got >= need:
        return {"ok": False, "verdict": "full", "need": need, "got": got,
                "name": item.get("name", "")}

    # Бутылка уже в реестре — её записали раньше. Это не придирка: без такой
    # проверки бутылку с собственной полки можно «принять» второй раз.
    old = await db.qr_get(code)
    if old:
        return {"ok": False, "verdict": "known", "name": item.get("name", ""),
                "label": old.get("label") or "",
                "district": old.get("district") or "",
                "at": str(old.get("at") or ""), "need": need, "got": got}

    now = datetime.now(timezone.utc)
    await db.supply_task_start(sid, oid, now)
    upd = await db.supply_take(sid, oid, pid, need, now)
    if not upd:
        return {"ok": False, "verdict": "full", "need": need, "got": got,
                "name": item.get("name", "")}

    from qr_routes import product_slug
    seq = await db.qr_next_seq(pid)
    label = f"{product_slug(pid, item.get('name',''))}#{seq:06d}"
    added = await db.qr_add(code, pid, item.get("name", ""), oid, tg_id, now, label,
                            extra={"src": "intake", "supply_id": sid, "driver": me,
                                   "at_dev": str(at_dev or "")[:32]})
    if not added:
        # Код заняли между проверкой и вставкой — место в задаче возвращаем,
        # иначе строка закроется бутылкой, которой у нас нет.
        await db.supply_untake(sid, oid, pid)
        return {"ok": False, "verdict": "known", "name": item.get("name", "")}

    # Бутылка встала на полку — заявка про неё ещё не знает.
    try:
        import stock_routes
        stock_routes.base_drop()
    except Exception:
        pass

    # ── то, что нельзя запретить, но нужно показать ──────────────────────────
    flags = []
    prev = task.get("last_at")
    if prev:
        try:
            ms = (now - prev).total_seconds() * 1000
            if 0 <= ms < FAST_MS:
                await db.supply_task_flag(sid, oid, {"kind": "fast_tick", "at": now})
                was = sum(1 for f in (task.get("flags") or [])
                          if f.get("kind") == "fast_tick")
                if was + 1 == FAST_LIMIT:
                    flags.append("fast")
        except Exception:
            pass
    sigs = await _known_sigs(pid)
    if len(sigs) >= 4 and _sig(code) not in sigs:
        flags.append("shape")
        await db.supply_task_flag(sid, oid, {"kind": "shape", "code": code[:40],
                                             "product": item.get("name", ""), "at": now})

    line = next((i for i in (upd.get("items") or []) if i["id"] == pid), {})
    got = int((line.get("got") or {}).get(oid) or got + 1)
    t = (upd.get("tasks") or {}).get(oid) or {}
    return {"ok": True, "verdict": "taken", "label": label, "code": code,
            "name": item.get("name", ""), "need": need, "got": got,
            "left": max(0, need - got), "flags": flags,
            "task_got": int(t.get("scanned") or 0)}


async def task_undo(sid: str, oid: str, code: str, me: str) -> dict:
    """Убрать последнюю бутылку — навёл камеру не на ту.

    Только свежую и только несколько раз за задачу: отмена без ограничений
    даёт ровно ту дыру, ради которой всё остальное и делалось."""
    sup = await db.supply_get(sid)
    if not sup:
        return {"ok": False, "verdict": "no_supply"}
    task = (sup.get("tasks") or {}).get(oid) or {}
    if task.get("driver") != me or task.get("done_at"):
        return {"ok": False, "verdict": "not_mine"}
    if int(task.get("undo") or 0) >= UNDO_MAX:
        return {"ok": False, "verdict": "undo_limit", "limit": UNDO_MAX}
    doc = await db.qr_get(code)
    if not doc or doc.get("supply_id") != sid:
        return {"ok": False, "verdict": "not_ours"}
    try:
        age = (datetime.now(timezone.utc) - doc["at"]).total_seconds()
    except Exception:
        age = 0
    if age > UNDO_SEC:
        return {"ok": False, "verdict": "too_late", "sec": UNDO_SEC}
    await db.qr_remove(code)
    await db.supply_untake(sid, oid, doc.get("product_id") or "")
    await db.supply_task_flag(sid, oid, {"kind": "undo", "code": code[:40],
                                         "product": doc.get("product_name", ""),
                                         "at": datetime.now(timezone.utc)})
    log.info(f"[supply] {sid}/{oid}: {me} отменил {code[:30]}")
    return {"ok": True, "code": code}


async def task_finish(sid: str, oid: str, me: str, note: str = "") -> dict:
    """Закрыть задачу района. Недобор считается сам — по строкам.

    Спрашивать «сколько не хватило» отдельно незачем: разницу между
    подтверждённым и принятым видно и без водителя. У него спрашиваем только
    то, чего в базе нет, — почему."""
    sup = await db.supply_get(sid)
    if not sup:
        return {"ok": False, "verdict": "no_supply"}
    task = (sup.get("tasks") or {}).get(oid) or {}
    if task.get("driver") != me:
        return {"ok": False, "verdict": "not_mine"}
    if task.get("done_at"):
        return {"ok": False, "verdict": "closed"}

    gaps = []
    for it in sup.get("items") or []:
        need = int((it.get("by_district") or {}).get(oid) or 0)
        if not need:
            continue
        got = int((it.get("got") or {}).get(oid) or 0)
        if got < need:
            gaps.append({"id": it["id"], "name": it.get("name", ""),
                         "need": need, "got": got, "gap": need - got})
    now = datetime.now(timezone.utc)
    doc = await db.supply_task_finish(sid, oid, gaps, str(note or "")[:300], now)
    if not doc:
        return {"ok": False, "verdict": "closed"}
    log.info(f"[supply] {sid}/{oid}: приёмка закрыта · {me} · "
             f"{(doc.get('tasks') or {}).get(oid, {}).get('scanned', 0)} шт · "
             f"недобор {sum(g['gap'] for g in gaps)}")
    try:
        await _notify_done(sid, doc, oid, me, gaps, note)
    except Exception as e:
        log.error(f"[supply] уведомление о приёмке: {e}")
    return {"ok": True, "gaps": gaps,
            "supply_done": doc.get("status") == "done"}


def _md(s: str) -> str:
    return str(s or "").replace("*", "").replace("_", "").replace("`", "")


async def _notify_done(sid: str, doc: dict, oid: str, me: str,
                       gaps: list, note: str):
    """Старшему — итог приёмки. Одно сообщение на район, а не на бутылку."""
    from owner_routes import notify_owners, notify_owners_force
    task = (doc.get("tasks") or {}).get(oid) or {}
    took = int(task.get("scanned") or 0)
    need = took + sum(g["gap"] for g in gaps)
    where = f"{OFFICE_CODES.get(oid,'')} {OFFICE_NAMES.get(oid, oid)}".strip()
    mins = ""
    try:
        if task.get("started_at"):
            mins = f" · {int((task['done_at'] - task['started_at']).total_seconds() // 60)} мин"
    except Exception:
        pass
    head = f"*Приёмка — {_md(where)}*\n{_md(me)} · принято {took} из {need}{mins}"
    if gaps:
        lst = "\n".join(f"• {_md(g['name'])} — {g['got']} из {g['need']}" for g in gaps[:8])
        more = f"\n…и ещё {len(gaps) - 8}" if len(gaps) > 8 else ""
        body = f"{head}\n\n*Недобор {sum(g['gap'] for g in gaps)} бутылок:*\n{lst}{more}"
        if note:
            body += f"\n\n_{_md(note)}_"
        await notify_owners("supply.done", body)
    else:
        await notify_owners("supply.done", head + "\n\nВзято полностью.")

    # Странности идут отдельным сообщением и мимо настроек: это не сводка, а
    # повод посмотреть записи приёмки своими глазами.
    fl = task.get("flags") or []
    fast = sum(1 for f in fl if f.get("kind") == "fast_tick")
    shape = [f for f in fl if f.get("kind") == "shape"]
    undo = sum(1 for f in fl if f.get("kind") == "undo")
    parts = []
    if fast >= FAST_LIMIT:
        parts.append(f"• {fast} сканов подряд быстрее {FAST_MS/1000:g} с")
    if shape:
        parts.append(f"• {len(shape)} кодов не похожи на остальные по этой позиции")
    if undo:
        parts.append(f"• отмен: {undo}")
    if parts:
        await notify_owners_force(
            "supply.flag",
            f"*Приёмка требует взгляда — {_md(where)}*\n{_md(me)}\n" + "\n".join(parts))


# ── чего не хватает: то, что придётся брать не здесь ────────────────────────
# Три разные вещи с одним следствием. Магазин отказал совсем; магазин урезал;
# магазин подтвердил, но при отгрузке не дал. Для полки это одно и то же —
# бутылок нет, — но разговор с магазином по ним разный, поэтому причина едет
# рядом со строкой, а не растворяется в общей сумме.
SHORT_WHY = {"dropped": "магазин отказал", "short": "магазин урезал",
             "gap": "не выдал при отгрузке"}


def _shortfall(sup: dict) -> dict:
    rows = {}

    def put(pid, name, why, total, by=None):
        """total — сколько не хватает, by — куда именно, если это известно.

        Разбивка по районам есть не всегда: у заявок, выгруженных до появления
        снимка по точкам, её взять неоткуда. Терять из-за этого саму строку
        нельзя — «не хватает семи бутылок» полезно и без адреса."""
        r = rows.setdefault(pid, {"id": pid, "name": name, "why": why,
                                  "by_district": {}, "gap": 0})
        r["gap"] += max(0, int(total or 0))
        for oid, n in (by or {}).items():
            n = int(n or 0)
            if n > 0:
                r["by_district"][oid] = r["by_district"].get(oid, 0) + n
        # Позиция может недобрать дважды: магазин урезал, а остаток ещё и не
        # выдал. Причину показываем последнюю — она ближе к делу.
        if why == "gap":
            r["why"] = why

    for d in sup.get("dropped") or []:
        put(d["id"], d.get("name", ""), "dropped", d.get("asked") or 0,
            d.get("by_district"))
    for s in sup.get("short") or []:
        put(s["id"], s.get("name", ""), "short", s.get("gap") or 0,
            s.get("by_district"))
    for oid, t in (sup.get("tasks") or {}).items():
        for g in t.get("gaps") or []:
            put(g["id"], g.get("name", ""), "gap", g.get("gap") or 0,
                {oid: g.get("gap") or 0})

    out = [r for r in rows.values() if r["gap"] > 0]
    out.sort(key=lambda r: (-r["gap"], r["name"]))
    return {"rows": out, "qty": sum(r["gap"] for r in out)}


def _short_book(sup: dict, short: dict):
    """Недобор одним листом — его несут в другой магазин.

    Тот же вид, что и у заявки, но без просьбы править: это уже не переговоры,
    а список на закупку. Причина стоит колонкой — в другом магазине спросят,
    почему берём именно это."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    dist = [o for o in OFFICE_IDS
            if any(r["by_district"].get(o) for r in short["rows"])]
    wb = Workbook(); ws = wb.active; ws.title = "Shortfall"
    ws.sheet_view.showGridLines = False
    white = Font(bold=True, size=11, color="FFFFFFFF"); bold = Font(bold=True, size=11)
    head_fill = PatternFill("solid", fgColor="FF1F2A37")
    num_fill = PatternFill("solid", fgColor="FF000000")
    band = PatternFill("solid", fgColor="FFEEECE1")
    sum_fill = PatternFill("solid", fgColor="FFFFFF00")
    thin = Side(style="thin"); box = Border(left=thin, right=thin, top=thin, bottom=thin)
    mid = Alignment(horizontal="center", vertical="center")

    N, C, I, W, D0 = 2, 3, 4, 5, 6
    LAST = D0 + len(dist) - 1; TOT = LAST + 1
    ws.cell(row=1, column=N, value=f"AMBAR · shortfall · {sup.get('day') or ''}")
    ws.cell(row=1, column=N).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=N, end_row=1, end_column=I)
    ws.row_dimensions[1].height = 21.6

    head = ["№", CODE_COL, "Item", "Reason"] + \
           [f"{OFFICE_CODES.get(o,'')} {DIST_EN.get(o, OFFICE_NAMES.get(o,o))}" for o in dist] + \
           [TOTAL_COL]
    for i, title in enumerate(head):
        c = ws.cell(row=3, column=N + i, value=title)
        c.alignment = mid; c.border = box
        if N + i == N:      c.fill = num_fill;  c.font = white
        elif N + i == TOT:  c.fill = sum_fill;  c.font = bold
        else:               c.fill = head_fill; c.font = white

    for n, r in enumerate(short["rows"], 1):
        i = 3 + n
        stripe = (n % 2 == 0)
        ws.cell(row=i, column=N, value=n).alignment = mid
        ws.cell(row=i, column=C, value=r["id"]).alignment = mid
        nm = ws.cell(row=i, column=I, value=r["name"]); nm.font = bold
        ws.cell(row=i, column=W, value=SHORT_WHY.get(r["why"], ""))
        for k, o in enumerate(dist):
            v = r["by_district"].get(o) or None
            c = ws.cell(row=i, column=D0 + k, value=v)
            c.alignment = mid; c.number_format = ZERO_BLANK
        t = ws.cell(row=i, column=TOT, value=r["gap"])
        t.font = bold; t.alignment = mid; t.fill = sum_fill
        for col in range(N, TOT + 1):
            c = ws.cell(row=i, column=col); c.border = box
            if stripe and col != TOT:
                c.fill = band
        ws.row_dimensions[i].height = 17.4

    i = 4 + len(short["rows"])
    ws.cell(row=i, column=N, value="TOTAL")
    ws.merge_cells(start_row=i, start_column=N, end_row=i, end_column=W)
    for col in range(N, TOT + 1):
        c = ws.cell(row=i, column=col)
        if col >= D0:
            o = dist[col - D0] if col <= LAST else None
            c.value = (sum(r["by_district"].get(o, 0) for r in short["rows"]) if o
                       else short["qty"]) or None
            c.number_format = ZERO_BLANK
        c.fill = sum_fill; c.font = bold; c.alignment = mid; c.border = box

    _tail(ws, i, N)

    ws.column_dimensions["A"].width = 29.55
    ws.column_dimensions[get_column_letter(N)].width = 7.11
    ws.column_dimensions[get_column_letter(C)].width = 10
    ws.column_dimensions[get_column_letter(I)].width = 40
    ws.column_dimensions[get_column_letter(W)].width = 22
    for col in range(D0, LAST + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions[get_column_letter(TOT)].width = 18.44
    ws.freeze_panes = f"{get_column_letter(D0)}4"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read(), f"AMBAR-shortfall-{sup.get('day') or sup['_id']}.xlsx"


def _supply_view(sup: dict) -> dict:
    sid = sup.get("_id")
    tasks = []
    for oid, t in (sup.get("tasks") or {}).items():
        v = _task_view(sid, sup, oid, t)
        fl = t.get("flags") or []
        v["flags"] = {
            "fast": sum(1 for f in fl if f.get("kind") == "fast_tick"),
            "shape": sum(1 for f in fl if f.get("kind") == "shape"),
            "undo": sum(1 for f in fl if f.get("kind") == "undo"),
        }
        v.pop("mine", None)
        tasks.append(v)
    tasks.sort(key=lambda t: OFFICE_IDS.index(t["district"])
               if t["district"] in OFFICE_IDS else 99)
    short = _shortfall(sup)
    return {
        "supply_id": sid, "at": str(sup.get("at") or ""), "day": sup.get("day") or "",
        "status": sup.get("status") or "open",
        "total_qty": sup.get("total_qty", 0), "asked_qty": sup.get("asked_qty", 0),
        "took": sum(t["got"] for t in tasks),
        "dropped": sup.get("dropped") or [], "short": sup.get("short") or [],
        "extra": sup.get("extra") or [], "unknown": sup.get("unknown") or [],
        "shortfall": short, "tasks": tasks,
        "districts": {o: {"code": OFFICE_CODES.get(o, ""),
                          "name": OFFICE_NAMES.get(o, o)} for o in OFFICE_IDS},
    }


@require_owner
async def handle_one(request):
    sup = await db.supply_get(request.match_info.get("sid") or "")
    if not sup:
        return web.json_response({"error": "not_found"}, status=404, headers=CORS_HEADERS)
    return web.json_response(_supply_view(sup), headers=CORS_HEADERS,
                             dumps=lambda o: __import__("json").dumps(o, default=str))


@require_owner
async def handle_release(request):
    """Снять задачу с водителя: взял и не поехал.

    Не удаляем принятое — только освобождаем задачу. Бутылки, которые он уже
    отсканировал, физически существуют и в остатке остаются."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    sid = request.match_info.get("sid") or ""
    oid = str(body.get("district") or "").strip()
    ok = await db.supply_task_release(sid, oid)
    if ok:
        log.info(f"[supply] {sid}/{oid}: задача снята владельцем")
    return web.json_response({"ok": ok}, headers=CORS_HEADERS)


@require_owner
async def handle_short_export(request):
    sup = await db.supply_get(request.match_info.get("sid") or "")
    if not sup:
        return web.json_response({"error": "not_found"}, status=404, headers=CORS_HEADERS)
    short = _shortfall(sup)
    if not short["rows"]:
        return web.json_response({"error": "nothing_missing"}, status=400,
                                 headers=CORS_HEADERS)
    raw, name = _short_book(sup, short)
    return web.Response(body=raw, headers={
        **CORS_HEADERS,
        "Content-Type": "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet",
        "Content-Disposition": f'attachment; filename="{name}"'})


@require_owner
async def handle_short_send(request):
    sup = await db.supply_get(request.match_info.get("sid") or "")
    if not sup:
        return web.json_response({"error": "not_found"}, status=404, headers=CORS_HEADERS)
    short = _shortfall(sup)
    if not short["rows"]:
        return web.json_response({"error": "nothing_missing"}, status=400,
                                 headers=CORS_HEADERS)
    raw, name = _short_book(sup, short)
    ok, err = await _send_doc(request.get("owner_id") or 0, raw, name,
                              f"Заказать на доп. складах · {short['qty']} бутылок")
    if not ok:
        return web.json_response({"error": err}, status=502, headers=CORS_HEADERS)
    return web.json_response({"ok": True, "file": name, "qty": short["qty"]},
                             headers=CORS_HEADERS)


async def _send_doc(chat_id: int, raw: bytes, name: str, caption: str):
    """Файл владельцу в переписку — оттуда его пересылают дальше."""
    from api_server import _aiohttp
    from owner_routes import OWNER_BOT_TOKEN
    if not OWNER_BOT_TOKEN:
        return False, "no_bot"
    form = _aiohttp.FormData()
    form.add_field("chat_id", str(chat_id))
    form.add_field("caption", caption)
    form.add_field("document", raw, filename=name,
                   content_type="application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet")
    url = f"https://api.telegram.org/bot{OWNER_BOT_TOKEN}/sendDocument"
    try:
        async with _aiohttp.ClientSession(
                timeout=_aiohttp.ClientTimeout(total=30)) as sess:
            async with sess.post(url, data=form) as r:
                res = await r.json()
    except Exception as e:
        log.error(f"[supply] отправка файла: {e}")
        return False, "send_failed"
    if not res.get("ok"):
        log.error(f"[supply] телеграм отказал: {res.get('description')}")
        return False, "telegram"
    return True, ""


async def _opt(request):
    return web.Response(status=200, headers=CORS_HEADERS)


def setup(app):
    r = app.router
    # Точные пути раньше шаблонных: aiohttp разбирает их в порядке добавления,
    # и «/supply/export» иначе попал бы в «/supply/{sid}» с sid=export.
    for path, handler, method in (
        ("/api/owner/supply/export", handle_export, "GET"),
        ("/api/owner/supply/send",   handle_send,   "POST"),
        ("/api/owner/supply/import", handle_import, "POST"),
        ("/api/owner/supply",        handle_list,   "GET"),
        ("/api/owner/supply/{sid}",                 handle_one,          "GET"),
        ("/api/owner/supply/{sid}/release",         handle_release,      "POST"),
        ("/api/owner/supply/{sid}/shortfall",       handle_short_export, "GET"),
        ("/api/owner/supply/{sid}/shortfall/send",  handle_short_send,   "POST"),
    ):
        r.add_route("OPTIONS", path, _opt)
        {"GET": r.add_get, "POST": r.add_post}[method](path, handler)
    log.info("[supply] routes mounted")
