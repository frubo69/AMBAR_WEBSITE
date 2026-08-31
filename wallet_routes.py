"""
AMBAR — кошелёк USDT: сколько лежит и что по нему ходило.

Кошелёк у нас уже был, но виден он был только изнутри одного вопроса: watcher
спрашивал «пришёл ли платёж по счёту №такому-то», и всё, что не относилось к
заказу, для приложения не существовало. Перевод мимо заказа, уход с кошелька,
сам остаток — про это нельзя было узнать, не открыв блокчейн-обозреватель.

Здесь тот же ключ TronGrid и те же права: только чтение. Приватных ключей у
сервера нет и не будет — отправить отсюда ничего нельзя, и кнопки такой не
появится.

Поступления бывают двух видов, и это не про качество, а про путь. Заказ из
клиентского бота выставляет счёт — такой приход приложение узнаёт по txid.
А личный заказ оператор ведёт руками: скидывает номер кошелька, деньги падают
напрямую, счёта нет и связать их не с чем. Оба вида законны, и на экране они
так и называются: через приложение и не через приложение.
"""
import io
import logging
import time as _t
from datetime import datetime, timedelta, timezone

DUBAI = timezone(timedelta(hours=4))

from aiohttp import web

import db
import tron
from config import TRON_RECEIVE_ADDRESS
from owner_auth import require_owner, CORS_HEADERS

log = logging.getLogger("wallet")

# Блокчейн не отвечает мгновенно, а на экран смотрят и листают его туда-сюда.
# Полминуты — это дешевле для ключа и незаметно глазу: подтверждение перевода
# всё равно занимает минуты.
TTL = 30
_CACHE: dict = {"at": 0.0, "data": None}


def _short(a: str) -> str:
    a = str(a or "")
    return f"{a[:6]}…{a[-4:]}" if len(a) > 12 else a


async def _build() -> dict:
    balance = await tron.get_balance(TRON_RECEIVE_ADDRESS)
    transfers = await tron.get_transfers(TRON_RECEIVE_ADDRESS)
    # None и пустой список — разные ответы: первое значит «не дозвонились», и
    # говорить в этом случае «переводов нет» — врать.
    offline = transfers is None
    transfers = transfers or []

    # Через приложение или напрямую: счёт знает txid своего перевода. А то, что
    # пришло напрямую, к заказу привязывает человек — и эта связь лежит рядом.
    ids = [t["txid"] for t in transfers if t.get("txid")]
    byid, links = {}, {}
    try:
        byid = await db.crypto_invoices_by_txids(ids)
    except Exception as e:                       # noqa: BLE001
        log.warning(f"[wallet] счета к переводам не подшились: {e}")
    try:
        links = await db.wallet_links(ids)
    except Exception as e:                       # noqa: BLE001
        log.warning(f"[wallet] привязки не подшились: {e}")

    через = 0.0
    напрямую = 0.0
    привязано = 0.0
    ушло = 0.0
    rows = []
    for t in transfers:
        tx = t.get("txid") or ""
        inv = byid.get(tx)
        link = links.get(tx)
        if t.get("in"):
            if inv:
                через += t["amount"]
            else:
                напрямую += t["amount"]
                if link:
                    привязано += t["amount"]
        else:
            ушло += t["amount"]
        rows.append({
            **t,
            "peer": _short(t.get("peer") or ""),
            "order_id": (inv or {}).get("order_id") or (link or {}).get("order_id") or "",
            # Своим счётом или рукой человека — разное знание, и путать их
            # нельзя: одно проверено блокчейном, второе — чьим-то решением.
            "linked": bool(link and not inv),
            "linked_by": (link or {}).get("by_name") or "",
        })

    # Сколько всего оплачено криптой по нашим счетам. Отдельно от ленты и
    # нарочно: лента — это последние переводы кошелька, а вопрос «сколько
    # прошло через приложение» про всю историю, и ответ на него лежит у нас, а
    # не в блокчейне.
    paid = {}
    try:
        paid = await db.crypto_paid_totals()
    except Exception as e:                       # noqa: BLE001
        log.warning(f"[wallet] итог по счетам не посчитан: {e}")

    # Строка в журнал: единственный способ проверить эти числа, не влезая в
    # чужой экран. Сходится ли остаток с балансом — видно сразу.
    log.info(f"[wallet] переводов {len(rows)} · пришло {round(через + напрямую, 2)} "
             f"(через приложение {round(через, 2)}) · ушло {round(ушло, 2)} · "
             f"остаток {round(через + напрямую - ушло, 2)} · "
             f"баланс {(balance or {}).get('usdt')}")

    return {
        "paid": paid,
        "address": _short(TRON_RECEIVE_ADDRESS),
        "balance": balance or {"usdt": 0.0, "trx": 0.0, "unknown": True},
        "offline": offline or balance is None,
        # Лента отдаётся целиком: обрезка на полусотне превращала историю в
        # «последние две недели», и человек справедливо спрашивал, где
        # остальное. Разворачивает её экран, по кнопке.
        "transfers": rows[:1000],
        "more": len(rows) > 1000,
        # «Ушло» тут не ради полноты: без него ноль на балансе выглядит
        # поломкой экрана, хотя это обычная жизнь кошелька — пришло и вывели.
        "totals": {"app": round(через, 2), "direct": round(напрямую, 2),
                   "linked": round(привязано, 2),
                   "in": round(через + напрямую, 2), "out": round(ушло, 2),
                   "n": len(rows)},
        "at": int(_t.time() * 1000),
    }


@require_owner
async def handle_wallet(request):
    """Баланс и последние переводы. Ответ держим полминуты на всех сразу."""
    if not TRON_RECEIVE_ADDRESS:
        return web.json_response({"error": "no_wallet"}, status=404, headers=CORS_HEADERS)
    свежий = str(request.query.get("fresh") or "") in ("1", "true")
    if not свежий and _CACHE["data"] and _t.monotonic() - _CACHE["at"] < TTL:
        return web.json_response({**_CACHE["data"], "cached": True},
                                 headers=CORS_HEADERS)
    data = await _build()
    # Ответ, в котором не дозвонились до сети, не кэшируем: следующее открытие
    # экрана должно попробовать заново, а не показывать ту же пустоту.
    if not data["offline"]:
        _CACHE.update(at=_t.monotonic(), data=data)
    return web.json_response(data, headers=CORS_HEADERS)


def _drop_cache():
    """Экран после привязки обязан показать её сразу, а не через полминуты."""
    _CACHE.update(at=0.0, data=None)


async def _link_note(kind: str, txid: str, oid: str, amount, who: str) -> None:
    try:
        from owner_routes import notify_owners_force, _md
    except Exception as e:                       # noqa: BLE001
        log.error(f"[wallet] письмо о привязке не отправлено: {e}")
        return
    шапка = ("🔗 *Поступление привязано к заказу*" if kind == "link"
             else "🔗 *Привязка поступления снята*")
    строки = [шапка, _md(who or "—"), "",
              f"• Заказ #{_md(oid)} — {amount} USDT",
              f"`{_md(str(txid)[:32])}`"]
    try:
        await notify_owners_force("wallet.link", "\n".join(строки))
    except Exception as e:                       # noqa: BLE001
        log.error(f"[wallet] письмо о привязке не ушло: {e}")


@require_owner
async def handle_link(request):
    """Связать прямое поступление с заказом.

    Только то, что пришло без счёта: у оплаты из бота связь уже есть, и
    перебивать её рукой значит спорить с блокчейном."""
    try:
        body = await request.json()
    except Exception:
        return web.json_response({"error": "invalid_json"}, status=400, headers=CORS_HEADERS)
    txid = str(body.get("txid") or "").strip()[:80]
    oid = str(body.get("order_id") or "").strip()[:40]
    if not txid or not oid:
        return web.json_response({"error": "txid_and_order_required"}, status=400,
                                 headers=CORS_HEADERS)
    order = await db.get_order(oid)
    if not order:
        return web.json_response({"error": "unknown_order"}, status=404, headers=CORS_HEADERS)
    inv = await db.crypto_invoices_by_txids([txid])
    if inv:
        return web.json_response({"error": "already_paid_by_invoice"}, status=409,
                                 headers=CORS_HEADERS)
    who = str(body.get("as") or "").strip()[:40]
    ok = await db.wallet_link_set(txid, {
        "order_id": oid, "amount": float(body.get("amount") or 0),
        "by": int(request.get("owner_id") or 0), "by_name": who,
        "at": datetime.now(timezone.utc)})
    if not ok:
        return web.json_response({"error": "not_saved"}, status=500, headers=CORS_HEADERS)
    _drop_cache()
    log.info(f"[wallet] привязка {txid[:16]}… → заказ {oid}")
    await _link_note("link", txid, oid, body.get("amount") or 0, who)
    return web.json_response({"ok": True}, headers=CORS_HEADERS)


@require_owner
async def handle_unlink(request):
    try:
        body = await request.json()
    except Exception:
        return web.json_response({"error": "invalid_json"}, status=400, headers=CORS_HEADERS)
    txid = str(body.get("txid") or "").strip()[:80]
    было = (await db.wallet_links([txid])).get(txid) or {}
    ok = await db.wallet_link_del(txid)
    _drop_cache()
    if ok:
        log.info(f"[wallet] привязка снята {txid[:16]}…")
        await _link_note("unlink", txid, было.get("order_id") or "—",
                         было.get("amount") or 0,
                         str(body.get("as") or "").strip()[:40])
    return web.json_response({"ok": ok}, headers=CORS_HEADERS)


# ── сопоставление с заказами ────────────────────────────────────────────────
# Прямое поступление можно узнать по сумме: заказ стоит N дирхам, а по нашему
# курсу это ровно столько-то USDT. Совпало число и совпал день — почти наверняка
# это он и есть.
#
# «Почти» здесь и есть вся суть, поэтому сопоставление НЕ привязывает молча. Оно
# показывает, что нашло, и привязывает только то, у чего нет второго кандидата:
# два заказа на одну сумму в один день — это не совпадение, а монетка, и решать
# такое должен человек.
MATCH_DAYS = 2            # на сколько дней вокруг перевода ищем заказ
MATCH_TOL = 0.02          # допуск по сумме: 2 % или 1 USDT, что больше
DEAD = ("cancelled", "declined", "canceled")


def _usdt_of(total_aed) -> float:
    from config import CRYPTO_AED_PER_USDT
    rate = CRYPTO_AED_PER_USDT or 3.5
    return round(float(total_aed or 0) / rate, 2)


async def _match(apply: bool, who: str = "") -> dict:
    """Свести прямые поступления с заказами по сумме и дню."""
    from datetime import timedelta
    transfers = await tron.get_transfers(TRON_RECEIVE_ADDRESS)
    if transfers is None:
        return {"error": "offline"}
    ids = [t["txid"] for t in transfers if t.get("txid")]
    byid = await db.crypto_invoices_by_txids(ids)
    links = await db.wallet_links()
    # Кандидаты — только безымянные приходы: у оплаты по счёту связь уже есть.
    сироты = [t for t in transfers
              if t.get("in") and t.get("txid")
              and t["txid"] not in byid and t["txid"] not in links]
    if not сироты:
        return {"pairs": [], "orphans": 0, "taken": 0, "ambiguous": 0}

    lo = min(t["ts"] for t in сироты) - MATCH_DAYS * 86400_000
    hi = max(t["ts"] for t in сироты) + MATCH_DAYS * 86400_000
    start = datetime.fromtimestamp(lo / 1000, timezone.utc).isoformat()
    end = datetime.fromtimestamp(hi / 1000, timezone.utc).isoformat()
    orders = await db.get_orders_in_range(
        start, end, limit=None,
        fields=["order_id", "timestamp", "total", "status", "customer_name",
                "payment_method", "paid"])
    # Заказ, за который уже заплатили криптой по счёту, второй раз не платят.
    занятые = {(v or {}).get("order_id") for v in byid.values()}
    занятые |= {(v or {}).get("order_id") for v in links.values()}
    свободные = [o for o in orders
                 if str(o.get("order_id") or "") not in занятые
                 and (o.get("status") or "") not in DEAD
                 and float(o.get("total") or 0) > 0]

    pairs, взято, спорных = [], 0, 0
    занято_сейчас = set()
    for t in sorted(сироты, key=lambda x: x["ts"]):
        нужно = t["amount"]
        допуск = max(1.0, нужно * MATCH_TOL)
        рядом = []
        for o in свободные:
            oid = str(o.get("order_id") or "")
            if not oid or oid in занято_сейчас:
                continue
            try:
                ts = datetime.fromisoformat(
                    str(o.get("timestamp") or "").replace("Z", "+00:00"))
            except ValueError:
                continue
            разрыв = abs((ts.timestamp() * 1000) - t["ts"]) / 3600_000
            if разрыв > MATCH_DAYS * 24:
                continue
            ожидали = _usdt_of(o.get("total"))
            расхождение = abs(ожидали - нужно)
            if расхождение <= допуск:
                рядом.append((расхождение, разрыв, oid, o, ожидали))
        рядом.sort(key=lambda x: (round(x[0], 2), x[1]))
        if not рядом:
            continue
        # Второй кандидат с той же суммой — это монетка, а не совпадение.
        спорно = len(рядом) > 1 and round(рядом[1][0], 2) == round(рядом[0][0], 2)
        расхождение, разрыв, oid, o, ожидали = рядом[0]
        if спорно:
            спорных += 1
        pairs.append({
            "txid": t["txid"], "amount": t["amount"], "ts": t["ts"],
            "order_id": oid, "order_total": int(o.get("total") or 0),
            "order_usdt": ожидали, "order_ts": str(o.get("timestamp") or ""),
            "name": o.get("customer_name") or "",
            "gap_h": round(разрыв, 1), "off": round(расхождение, 2),
            "sure": not спорно,
            "others": len(рядом) - 1,
        })
        if спорно:
            continue
        занято_сейчас.add(oid)
        if apply:
            await db.wallet_link_set(t["txid"], {
                "order_id": oid, "amount": t["amount"],
                "by": 0, "by_name": who or "сопоставление",
                "auto": True, "at": datetime.now(timezone.utc)})
            взято += 1
    if apply:
        _drop_cache()
        log.info(f"[wallet] сопоставление: привязано {взято} из {len(сироты)}")
    return {"pairs": pairs, "orphans": len(сироты), "taken": взято,
            "ambiguous": спорных}


@require_owner
async def handle_match(request):
    """GET — что нашлось, POST — привязать найденное без спорных."""
    apply = request.method == "POST"
    who = ""
    if apply:
        try:
            who = str((await request.json()).get("as") or "").strip()[:40]
        except Exception:
            who = ""
    out = await _match(apply, who)
    if out.get("error") == "offline":
        return web.json_response(out, status=503, headers=CORS_HEADERS)
    if apply and out.get("taken"):
        try:
            from owner_routes import notify_owners_force, _md
            await notify_owners_force(
                "wallet.link",
                "🔗 *Поступления сведены с заказами*\n"
                + _md(who or "—") + "\n\n"
                + f"• привязано {out['taken']} "
                + ("· спорных оставлено " + str(out["ambiguous"])
                   if out.get("ambiguous") else "· спорных нет"))
        except Exception as e:                   # noqa: BLE001
            log.error(f"[wallet] письмо о сопоставлении не ушло: {e}")
    return web.json_response(out, headers=CORS_HEADERS)


# ── файлом в чат ────────────────────────────────────────────────────────────
# Список переводов на экране — чтобы посмотреть, а файл — чтобы работать: свести
# с бухгалтерией, отправить дальше, оставить у себя. Скачать из мини-приложения
# некуда, поэтому бот кладёт документ владельцу в переписку, как и заявку.


def _book(data: dict, only_linked: bool = False):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    rows = [r for r in (data.get("transfers") or [])
            if not only_linked or r.get("order_id")]
    wb = Workbook()
    ws = wb.active
    ws.title = "Переводы"
    шапка = ["Дата", "Время", "Направление", "USDT", "Заказ", "Как связано",
             "Вторая сторона", "Транзакция"]
    ws.append(шапка)
    for i, _ in enumerate(шапка, 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A1A32")
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        д = datetime.fromtimestamp((r.get("ts") or 0) / 1000, timezone.utc)
        д = д.astimezone(DUBAI)
        связь = ("счёт из бота" if r.get("order_id") and not r.get("linked")
                 else "сверка / вручную" if r.get("linked") else "")
        ws.append([д.strftime("%d.%m.%Y"), д.strftime("%H:%M"),
                   "приход" if r.get("in") else "расход",
                   round(float(r.get("amount") or 0), 2),
                   str(r.get("order_id") or ""), связь,
                   str(r.get("peer") or ""), str(r.get("txid") or "")])
    for кол, ширина in zip("ABCDEFGH", (12, 8, 13, 12, 16, 18, 20, 46)):
        ws.column_dimensions[кол].width = ширина
    ws.freeze_panes = "A2"

    итог = data.get("totals") or {}
    ws.append([])
    ws.append(["Пришло всего", "", "", итог.get("in")])
    ws.append(["Ушло с кошелька", "", "", итог.get("out")])
    ws.append(["Остаток", "", "", round((итог.get("in") or 0) - (итог.get("out") or 0), 2)])
    ws.append(["Оплачено по нашим счетам", "", "", (data.get("paid") or {}).get("usdt")])
    for i in range(len(rows) + 3, len(rows) + 7):
        ws.cell(row=i, column=1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(rows)


@require_owner
async def handle_export(request):
    """Прислать переводы файлом в чат с ботом."""
    from api_server import _aiohttp
    from owner_routes import OWNER_BOT_TOKEN
    if not OWNER_BOT_TOKEN:
        return web.json_response({"error": "no_bot"}, status=500, headers=CORS_HEADERS)
    only = str(request.query.get("linked") or "") in ("1", "true")
    data = _CACHE["data"] if _CACHE["data"] else await _build()
    raw, n = _book(data, only)
    имя = ("ambar-postupleniya" if only else "ambar-koshelek") + \
          f"-{datetime.now(DUBAI).strftime('%Y%m%d')}.xlsx"
    итог = data.get("totals") or {}
    подпись = (f"Поступления, сведённые с заказами · {n}" if only else
               f"Кошелёк USDT · {n} "
               + ("перевод" if n % 10 == 1 and n % 100 != 11 else "переводов")
               + f"\nПришло {итог.get('in')} · ушло {итог.get('out')} · "
               + f"остаток {round((итог.get('in') or 0) - (итог.get('out') or 0), 2)}")
    form = _aiohttp.FormData()
    form.add_field("chat_id", str(request.get("owner_id") or 0))
    form.add_field("caption", подпись)
    form.add_field("document", raw, filename=имя,
                   content_type="application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet")
    url = f"https://api.telegram.org/bot{OWNER_BOT_TOKEN}/sendDocument"
    to = _aiohttp.ClientTimeout(total=30)
    try:
        async with _aiohttp.ClientSession(timeout=to) as sess:
            async with sess.post(url, data=form) as r:
                res = await r.json()
    except Exception as e:                       # noqa: BLE001
        log.error(f"[wallet] отправка файла: {e}")
        return web.json_response({"error": "send_failed"}, status=502, headers=CORS_HEADERS)
    if not res.get("ok"):
        log.error(f"[wallet] телеграм отказал: {res.get('description')}")
        return web.json_response({"error": "telegram"}, status=502, headers=CORS_HEADERS)
    # В реестр переписки владельца: по тревоге файл должен уходить вместе со
    # всеми — в нём номера транзакций и суммы.
    try:
        from api_server import _remember_owner_msg
        await _remember_owner_msg(OWNER_BOT_TOKEN, request.get("owner_id") or 0, res)
    except Exception as e:                       # noqa: BLE001
        log.debug(f"[wallet] реестр файла: {e}")
    return web.json_response({"ok": True, "rows": n}, headers=CORS_HEADERS)


async def _opt(request):
    return web.Response(status=200, headers=CORS_HEADERS)


def setup(app):
    app.router.add_route("OPTIONS", "/api/owner/wallet", _opt)
    app.router.add_get("/api/owner/wallet", handle_wallet)
    app.router.add_route("OPTIONS", "/api/owner/wallet/link", _opt)
    app.router.add_post("/api/owner/wallet/link", handle_link)
    app.router.add_route("OPTIONS", "/api/owner/wallet/unlink", _opt)
    app.router.add_post("/api/owner/wallet/unlink", handle_unlink)
    app.router.add_route("OPTIONS", "/api/owner/wallet/export", _opt)
    app.router.add_post("/api/owner/wallet/export", handle_export)
    app.router.add_route("OPTIONS", "/api/owner/wallet/match", _opt)
    app.router.add_get("/api/owner/wallet/match", handle_match)
    app.router.add_post("/api/owner/wallet/match", handle_match)
    log.info("[wallet] routes mounted")
